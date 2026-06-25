"""
scraper.py — Station Austin Mentor Profile Scraper
----------------------------------------------------
Reads mentor profile URLs from `urls.txt` (one per line), scrapes each page,
and writes a professionally formatted `mentors.xlsx` Excel file.

Columns: Name | Company | Designation | Tags | LinkedIn | Mentor URL

Usage:
    python scraper.py

Requirements:
    pip install requests beautifulsoup4 openpyxl lxml
"""

import re
import sys
import time
import requests
from bs4 import BeautifulSoup, NavigableString
from urllib.parse import urlparse, urlencode, parse_qs, urlunparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  Configuration
# ─────────────────────────────────────────────
INPUT_FILE    = "urls.txt"
OUTPUT_FILE   = "mentors.xlsx"
REQUEST_DELAY = 0.8      # polite delay (seconds) between requests
TIMEOUT       = 20       # HTTP request timeout
MAX_RETRIES   = 2        # retry failed pages up to this many times

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

LINKEDIN_TRACKING_PARAMS = {
    "lipi", "trackingId", "trk", "trkInfo", "originalSubdomain",
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "rcm", "src",
}

# Placeholder values that mean "no company set" on this WordPress site
EMPTY_COMPANY_PLACEHOLDERS = {"-", "–", "—", "/", ".", "N/A", "n/a", "na", ""}

COLUMNS = ["Name", "Company", "Designation", "Tags", "LinkedIn", "Mentor URL"]


# ─────────────────────────────────────────────
#  URL helpers
# ─────────────────────────────────────────────
def clean_linkedin_url(raw: str) -> str:
    """Normalise a LinkedIn /in/ profile URL — strip tracking params."""
    if not raw or "linkedin.com/in/" not in raw.lower():
        return ""
    parsed = urlparse(raw)
    qs = parse_qs(parsed.query, keep_blank_values=False)
    clean_qs = {k: v for k, v in qs.items() if k not in LINKEDIN_TRACKING_PARAMS}
    new_query = urlencode(clean_qs, doseq=True) if clean_qs else ""
    return urlunparse((
        parsed.scheme or "https", parsed.netloc,
        parsed.path.rstrip("/"), parsed.params, new_query, "",
    ))


# ─────────────────────────────────────────────
#  Page fetcher with retry
# ─────────────────────────────────────────────
def fetch_page(url: str, session: requests.Session) -> BeautifulSoup | None:
    for attempt in range(1, MAX_RETRIES + 2):
        try:
            resp = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except requests.RequestException as exc:
            if attempt <= MAX_RETRIES:
                wait = attempt * 2
                print(f"    ⚠  Retry {attempt}/{MAX_RETRIES} — {exc}")
                time.sleep(wait)
            else:
                print(f"    ✗  Failed after {MAX_RETRIES + 1} attempts: {exc}")
                return None


# ─────────────────────────────────────────────
#  Data extractor
# ─────────────────────────────────────────────
def _get_designation(h1) -> str:
    """More robust designation extractor."""
    if not h1:
        return ""

    selectors = [
        "p.text-uppercase",
        "p",
        "div p",
        "span",
    ]

    for sel in selectors:
        try:
            el = h1.find_next(sel.split()[-1])
            while el:
                text = el.get_text(" ", strip=True)
                if (
                    text
                    and len(text) < 120
                    and "about" not in text.lower()
                    and text.lower() != h1.get_text(strip=True).lower()
                ):
                    return text
                el = el.find_next(el.name)
        except Exception:
            pass

    for sib in h1.next_siblings:
        if isinstance(sib, NavigableString):
            txt = str(sib).strip()
            if txt:
                return txt

    return ""


def extract_mentor_data(soup: BeautifulSoup, mentor_url: str) -> dict:
    data = {col: "" for col in COLUMNS}
    data["Mentor URL"] = mentor_url

    # ── Name ──────────────────────────────────────────────────────────────────
    h1 = soup.find("h1")
    if h1:
        data["Name"] = h1.get_text(strip=True)

    # ── Designation ───────────────────────────────────────────────────────────
    data["Designation"] = _get_designation(h1)

    # ── Company ───────────────────────────────────────────────────────────────
    building_img = soup.find("img", src=re.compile(r"icon-building", re.I))
    if building_img:
        # Try direct next sibling (NavigableString " Company Name")
        company = ""
        for sib in building_img.next_siblings:
            sib_text = (
                str(sib).strip()
                if isinstance(sib, NavigableString)
                else sib.get_text(strip=True)
            )
            if sib_text:
                company = sib_text
                break
        # Fallback: all text in the parent element
        if not company:
            company = building_img.parent.get_text(strip=True)
        # Reject placeholders ("-", "—", etc.)
        if company.strip() in EMPTY_COMPANY_PLACEHOLDERS:
            company = ""
        data["Company"] = company

    # ── Tags ──────────────────────────────────────────────────────────────────
    tags = []
    hr_tag = soup.find("hr")
    if hr_tag:
        ul = hr_tag.find_next("ul")
        if ul:
            tags = [
                li.get_text(strip=True)
                for li in ul.find_all("li")
                if li.get_text(strip=True)
            ]
    # Fallback: tag pills shown above the h1
    if not tags and h1:
        prev = h1.find_previous_sibling()
        if prev:
            for c in prev.find_all(["span", "a", "li", "div"]):
                t = c.get_text(strip=True)
                if t and len(t) < 60:
                    tags.append(t)
    data["Tags"] = ", ".join(tags)

    # ── LinkedIn ──────────────────────────────────────────────────────────────
    # Strategy 1: anchor wrapping the LinkedIn icon image
    li_img = soup.find("img", src=re.compile(r"icon-linkedin", re.I))
    if li_img:
        anchor = li_img.find_parent("a")
        if anchor and anchor.get("href") and "linkedin.com/in/" in anchor["href"]:
            data["LinkedIn"] = clean_linkedin_url(anchor["href"])
    # Strategy 2: any <a> pointing to a personal linkedin.com/in/ profile
    if not data["LinkedIn"]:
        for anchor in soup.find_all("a", href=True):
            if "linkedin.com/in/" in anchor["href"]:
                cleaned = clean_linkedin_url(anchor["href"])
                if cleaned:
                    data["LinkedIn"] = cleaned
                    break

    return data


# ─────────────────────────────────────────────
#  Excel writer
# ─────────────────────────────────────────────
HEADER_BG    = "1F3864"
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
ROW_FONT     = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", fgColor="EEF2F7")
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN    = Alignment(vertical="top")
CELL_BORDER  = Border(
    bottom=Side(style="thin", color="D0D7E2"),
    right=Side(style="thin", color="D0D7E2"),
)
COL_WIDTHS   = {
    "Name": 28, "Company": 32, "Designation": 40,
    "Tags": 55, "LinkedIn": 50, "Mentor URL": 48,
}


def write_excel(records: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mentors"

    # Header row
    ws.append(COLUMNS)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = header_fill
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, record in enumerate(records, start=2):
        ws.append([record.get(col, "") for col in COLUMNS])
        row_fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font   = ROW_FONT
            cell.border = CELL_BORDER
            cell.alignment = WRAP_ALIGN if col_name in ("Tags", "Designation") else TOP_ALIGN
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 22

    # Hyperlinks for LinkedIn and Mentor URL
    li_col  = COLUMNS.index("LinkedIn") + 1
    url_col = COLUMNS.index("Mentor URL") + 1
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    for row_idx, record in enumerate(records, start=2):
        for col_idx, field in [(li_col, "LinkedIn"), (url_col, "Mentor URL")]:
            val = record.get(field, "")
            if val:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.hyperlink = val
                cell.font = link_font

    # Auto-fit column widths
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        min_w = COL_WIDTHS.get(col_name, 20)
        max_len = min_w
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[col_letter].width = max_len + 3

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────
def main():
    try:
        with open(INPUT_FILE, encoding="utf-8") as fh:
            urls = [line.strip() for line in fh if line.strip()]
    except FileNotFoundError:
        print(f"ERROR: '{INPUT_FILE}' not found.")
        sys.exit(1)

    if not urls:
        print(f"ERROR: '{INPUT_FILE}' is empty.")
        sys.exit(1)

    total = len(urls)
    print(f"\n{'─'*55}")
    print(f"  Station Austin Mentor Scraper  (v2 — fixed)")
    print(f"  {total} URLs loaded from '{INPUT_FILE}'")
    print(f"{'─'*55}\n")

    records, failed = [], []
    session = requests.Session()

    for idx, url in enumerate(urls, start=1):
        print(f"  [{idx:>3}/{total}] {url}")

        soup = fetch_page(url, session)
        if soup is None:
            failed.append(url)
            records.append({col: "" for col in COLUMNS} | {"Mentor URL": url})
            continue

        try:
            record = extract_mentor_data(soup, url)
        except Exception as exc:
            print(f"    ✗  Parse error: {exc}")
            failed.append(url)
            records.append({col: "" for col in COLUMNS} | {"Mentor URL": url})
            continue

        records.append(record)
        name  = record["Name"]        or "(no name)"
        desig = record["Designation"] or "(no designation)"
        comp  = record["Company"]     or "(no company)"
        print(f"         ✓  {name} | {desig} | {comp}")
        time.sleep(REQUEST_DELAY)

    print(f"\n{'─'*55}")
    print(f"  Writing {len(records)} records to '{OUTPUT_FILE}' …")
    write_excel(records, OUTPUT_FILE)
    print(f"  ✓  Saved: {OUTPUT_FILE}")

    success = total - len(failed)
    print(f"\n{'─'*55}")
    print(f"  DONE  ✓  {success}/{total} profiles scraped successfully")
    if failed:
        print(f"       ✗  {len(failed)} failed:")
        for f in failed:
            print(f"            • {f}")
    print(f"{'─'*55}\n")


if __name__ == "__main__":
    main()